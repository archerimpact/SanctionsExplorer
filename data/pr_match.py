import os
from openpyxl import load_workbook
from openpyxl import Workbook
import requests
import urllib
import json

url_template = "http://localhost:9200/pr/pr/_search"
headers = {'Content-Type': 'application/json'}
data = {"query": {"match_phrase": {"content" :{"query": "", "slop": "3"}}}}

#wb = load_workbook(filename='../xml/sdn.xlsx')
#ws = wb.active

jsondata = {}
#writefile = open("match_data.txt", "w")

scandata = {"_source": ["primary_display_name"], "size": "10000"}
scan = requests.get("http://localhost:9200/sdn/_search", json=scandata)
if scan.status_code == 200:
    print("scan worked")
    for elem in scan.json()["hits"]["hits"]:
        val = elem["_source"]["primary_display_name"]
        print(val)
        sdnid = elem["_id"]
        jsondata[sdnid] = []
        data["query"]["match_phrase"]["content"]["query"] = val;
        result = requests.get(url_template, json=data)
        if result.status_code == 200:
            for entry in result.json()["hits"]["hits"]:
                pr_elem = {
                    "pr_id": entry["_id"],
                    "link": entry["_source"]["link"],
                    "date": entry["_source"]["date"],
                    "title": entry["_source"]["title"].encode("utf-8")
                }
                jsondata[sdnid].append(pr_elem)
                #writefile.write(sdnid + " | ")
                #writefile.write(val.encode("utf-8"))
                #writefile.write(" | " + entry["_id"] + " | " + entry["_source"]["link"] + " | " + entry["_source"]["date"] + " | ")
                #writefile.write(entry["_source"]["title"].encode("utf-8"))
                #writefile.write("\n")
        else:
            print("fail")
    with open("entries.json", "w") as outfile:
        json.dump(jsondata, outfile)
else:
    print("scan failed")
#new_wb = Workbook()
#new_ws = new_wb.active
#rownum = 1
#x = 0
#for row in ws.rows:
#	val = row[1].value
#	print(val)
#	if val != None and isinstance(val, unicode):
#		url = url_template + urllib.quote_plus(val).replace("+", "%20")
		#print(url)
#		data["query"]["match_phrase"]["content"]["query"] = val;
		#print(urllib.urlencode(data))
#		result = requests.get(url_template,json=data)
                #print(result.url)
#		if result.status_code == 200:
                        #print(result.json())
#			for entry in result.json()["hits"]["hits"]:
                                #print entry
				#new_ws.cell(row=rownum, column=1).value = val
				#new_ws.cell(row=rownum, column=2).value = entry["link"]
                		#new_ws.cell(row=rownum, column=3).value = entry["date"]
                		#new_ws.cell(row=rownum, column=4).value = entry["title"]
				#print(entry["link"])
				#writefile.write(val + " | " + entry["link"] + " | " + entry["date"] + " | ")
				#writefile.write(entry["title"].encode("utf-8"))
                                #print entry.keys()
   #                             writefile.write(val + " | " + entry["_source"]["link"] + " | " + entry["_source"]["date"] + " | ")
  #                              writefile.write(entry["_source"]["title"].encode("utf-8"))
 #                               writefile.write("\n")
#				rownum += 1
#		else:
#			print(result.content)

#writefile.close()

#new_wb.save("prelim_matches.xlsx")
#result = requests.get(url)
#print(result.status_code)
#print(result.json()["response"][0].keys())		
