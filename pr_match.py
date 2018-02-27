import os
from openpyxl import load_workbook
from openpyxl import Workbook
import requests

url_template = "http://localhost:8080/search/press-releases?query="
headers = {'Content-Type': 'application/json'}
data = {"query": "Colima"}

wb = load_workbook(filename='sdn.xlsx')
ws = wb.active

new_wb = Workbook()
new_ws = new_wb.active
rownum = 1

for row in ws.rows:
	val = row[1].value
	print(type(val))
	print(val)
	if val != None and isinstance(val, unicode):
		url = url_template + val
		result = requests.get(url, headers=headers, data=data)
		if result.status_code == 200:
			for entry in result.json()["response"]:
				new_ws.cell(row=rownum, column=1).value = val
				new_ws.cell(row=rownum, column=2).value = entry["link"]
				print(entry["link"])
				rownum += 1
		else:
			print("error")

new_wb.save("prelim_matches.xlsx")
#result = requests.get(url)
#print(result.status_code)
#print(result.json()["response"][0].keys())		
